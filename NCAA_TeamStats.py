# -*- coding: utf-8 -*-
"""
Created on Tue Dec 24 14:25:35 2019

@author: kiran
"""
import openpyxl
#import pandas as pd
#from urllib.request import urlopen as uReq
import urllib
#from selenium import webdriver
from bs4 import BeautifulSoup as bs
import datetime
#from openpyxl.styles import PatternFill
date_object = datetime.date.today() + datetime.timedelta(days = 1)
print(date_object.strftime('%Y-%m-%d'))

scheduleurl = 'https://www.teamrankings.com/ncb/schedules/?date='+ date_object.strftime('%Y-%m-%d')

raw_html = urllib.request.urlopen(scheduleurl)
html = bs(raw_html,'html.parser')
#print(html)


try:
    Match = html.find_all('td')[2]
except IndexError:
    print('No matched scheduled for tomorrow.')           
whitelist = set('abcdefghijklmnopqrstuvwxyz ABCDEFGHIJKLMNOPQRSTUVWXYZ-()&.')
Matches = html.find_all('table')[0].find_all('td')
i = 3 
teamlist = []  
for mat in Matches:
    if i%5 == 0:
        text =  ''.join(filter(whitelist.__contains__, mat.get_text()))
        teams = text.split(' vs ')
        if len(teams) == 1:
            teams = teams[0].split(' at ')
        if len(teams) == 1:
            teams = text.split(' vs. ')
        
        print(teams)
        teamlist.append(teams[0])
        teamlist.append(teams[1])
    i = i + 1

print(teamlist)

Statswb = openpyxl.Workbook()
sheet1 = Statswb.active
sheet1.cell(1,1).value = 'Team Name'
sheet1.cell(1,2).value = 'Total Rebounding pct'
sheet1.cell(1,3).value = 'Free throw pct'
sheet1.cell(1,4).value = 'Schedule strength'
sheet1.cell(1,5).value = 'Points per Game'
sheet1.cell(1,6).value = 'Offensive Efficiency'
sheet1.cell(1,7).value = 'Defensive Efficiency'
sheet1.cell(1,8).value = 'Personal fouls per game'
sheet1.cell(1,9).value = 'Turnovers per game'
sheet1.cell(1,10).value = 'Opponent Turnovers per game'
sheet1.cell(1,11).value = '3 point pct'
sheet1.cell(1,12).value = 'Opponent 3 point pct'

k=2
for teamname in teamlist:
    sheet1.cell(k,1).value = teamname
    k = k + 1

stats_table1 = []
totalreboundurl = 'https://www.teamrankings.com/ncaa-basketball/stat/total-rebounding-percentage'
raw_html = urllib.request.urlopen(totalreboundurl)
html = bs(raw_html,'html.parser')
r = 2
for teamname in teamlist:

    
    Rankings = html.find_all('table')[0].find_all('tr')
    #print('Test' +teamname)         
    row_marker = 0
    row_data=[]  
    flag = 0
    for row in html.find_all('table')[0].find_all('tr'):
            column_marker = 0
            columns = row.find_all('td')
            row_data=[]
            for column in columns:
                if column_marker <= 1:
                    row_data.append(column.get_text())
                #print(row_data[column_marker])
                column_marker += 1                        
            for j in range(len(row_data)):
                #print('Test2'+teamname + ' '+ row_data[j] + ' '+ str(j) + str(j==1) + str(row_data[j].strip() == teamname.strip()))
                if j==1 and row_data[j].strip() == teamname.strip():
                    stats_table1.append(row_data)
                    cellx = sheet1.cell(row=r,column=2)
                    cellx.value = int(row_data[0].strip())
                    print(str(cellx.value) + ' - ' + teamname)
                    r = r + 1
                    flag = 1
    if flag == 0:
        cellx = sheet1.cell(row=r,column=2)
        cellx.value = 0
        r = r + 1
print('Finished...')        
#Statswb.save("D:\\upwork\\NCAA_Stats.xlsx")
#Statswb.close



stats_table2 = []
totalreboundurl = 'https://www.teamrankings.com/ncaa-basketball/stat/free-throw-pct'
raw_html = urllib.request.urlopen(totalreboundurl)
html = bs(raw_html,'html.parser')
r = 2
for teamname in teamlist:

    Rankings = html.find_all('table')[0].find_all('tr')
    #print('Test' +teamname)         
    row_marker = 0
    row_data=[]   
    flag = 0
    for row in html.find_all('table')[0].find_all('tr'):
            column_marker = 0
            columns = row.find_all('td')
            row_data=[]
            for column in columns:
                if column_marker <= 1:
                    row_data.append(column.get_text())
                #print(row_data[column_marker])
                column_marker += 1                        
            for j in range(len(row_data)):
                #print('Test2'+teamname + ' '+ row_data[j] + ' '+ str(j) + str(j==1) + str(row_data[j].strip() == teamname.strip()))
                if j==1 and row_data[j].strip() == teamname.strip():
                    stats_table2.append(row_data)
                    cellx = sheet1.cell(row=r,column=3)
                    cellx.value = int(row_data[0].strip())
                    print(str(cellx.value) + ' - ' + teamname)
                    r = r + 1
                    flag = 1
    if flag == 0:
        cellx = sheet1.cell(row=r,column=3)
        cellx.value = 0
        r = r + 1
print('Finished..')
stats_table3 = []
totalreboundurl = 'https://www.teamrankings.com/ncaa-basketball/ranking/schedule-strength-by-other'
raw_html = urllib.request.urlopen(totalreboundurl)
html = bs(raw_html,'html.parser')
r = 2
for teamname in teamlist:

    
    Rankings = html.find_all('table')[0].find_all('tr')
    #print('Test' +teamname)         
    row_marker = 0
    row_data=[]  
    flag = 0
    for row in html.find_all('table')[0].find_all('tr'):
            column_marker = 0
            columns = row.find_all('td')
            row_data=[]
            for column in columns:
                if column_marker <= 1:
                    if column_marker == 1:
                        row_data.append(column.find('a').get_text())
                    else:
                        row_data.append(column.get_text())
                #print(row_data[column_marker])
                column_marker += 1                        
            for j in range(len(row_data)):
                #print('Test2'+teamname + ' '+ row_data[j] + ' '+ str(j) + str(j==1) + str(row_data[j].strip() == teamname.strip()))
                #if j==1 and teamname.strip()+' (' in row_data[j].strip() :
                if j==1 and teamname.strip() == row_data[j].strip() :
                    stats_table3.append(row_data)
                    cellx = sheet1.cell(row=r,column=4)
                    cellx.value = int(row_data[0].strip())
                    print(str(cellx.value) + ' - ' + teamname)
                    r = r + 1
                    flag = 1
    if flag == 0:
        cellx = sheet1.cell(row=r,column=4)
        cellx.value = 0
        r = r + 1
        
stats_table4 = []
totalreboundurl = 'https://www.teamrankings.com/ncaa-basketball/stat/points-per-game'
raw_html = urllib.request.urlopen(totalreboundurl)
html = bs(raw_html,'html.parser')
r = 2
for teamname in teamlist:

    
    Rankings = html.find_all('table')[0].find_all('tr')
    #print('Test' +teamname)         
    row_marker = 0
    row_data=[]   
    flag = 0
    for row in html.find_all('table')[0].find_all('tr'):
            column_marker = 0
            columns = row.find_all('td')
            row_data=[]
            for column in columns:
                if column_marker <= 1:
                    row_data.append(column.get_text())
                #print(row_data[column_marker])
                column_marker += 1                        
            for j in range(len(row_data)):
                #print('Test2'+teamname + ' '+ row_data[j] + ' '+ str(j) + str(j==1) + str(row_data[j].strip() == teamname.strip()))
                if j==1 and row_data[j].strip() == teamname.strip():
                    stats_table4.append(row_data)
                    cellx = sheet1.cell(row=r,column=5)
                    cellx.value = int(row_data[0].strip())
                    print(str(cellx.value) + ' - ' + teamname)
                    r = r + 1
                    flag = 1
    if flag == 0:
        cellx = sheet1.cell(row=r,column=5)
        cellx.value = 0
        r = r + 1
    
stats_table5 = []
totalreboundurl = 'https://www.teamrankings.com/ncaa-basketball/stat/offensive-efficiency'
raw_html = urllib.request.urlopen(totalreboundurl)
html = bs(raw_html,'html.parser')
r = 2
for teamname in teamlist:

    
    Rankings = html.find_all('table')[0].find_all('tr')
    #print('Test' +teamname)         
    row_marker = 0
    row_data=[]  
    flag = 0
    for row in html.find_all('table')[0].find_all('tr'):
            column_marker = 0
            columns = row.find_all('td')
            row_data=[]
            for column in columns:
                if column_marker <= 1:
                    row_data.append(column.get_text())
                #print(row_data[column_marker])
                column_marker += 1                        
            for j in range(len(row_data)):
                #print('Test2'+teamname + ' '+ row_data[j] + ' '+ str(j) + str(j==1) + str(row_data[j].strip() == teamname.strip()))
                if j==1 and row_data[j].strip() == teamname.strip():
                    stats_table5.append(row_data)
                    cellx = sheet1.cell(row=r,column=6)
                    cellx.value = int(row_data[0].strip())
                    print(str(cellx.value) + ' - ' + teamname)
                    r = r + 1
                    flag = 1
    if flag == 0:
        cellx = sheet1.cell(row=r,column=6)
        cellx.value = 0
        r = r + 1
    
stats_table6 = []
totalreboundurl = 'https://www.teamrankings.com/ncaa-basketball/stat/defensive-efficiency'
raw_html = urllib.request.urlopen(totalreboundurl)
html = bs(raw_html,'html.parser')
r = 2
for teamname in teamlist:

    
    Rankings = html.find_all('table')[0].find_all('tr')
    #print('Test' +teamname)         
    row_marker = 0
    row_data=[]    
    flag = 0
    for row in html.find_all('table')[0].find_all('tr'):
            column_marker = 0
            columns = row.find_all('td')
            row_data=[]
            for column in columns:
                if column_marker <= 1:
                    row_data.append(column.get_text())
                #print(row_data[column_marker])
                column_marker += 1                        
            for j in range(len(row_data)):
                #print('Test2'+teamname + ' '+ row_data[j] + ' '+ str(j) + str(j==1) + str(row_data[j].strip() == teamname.strip()))
                if j==1 and row_data[j].strip() == teamname.strip():
                    stats_table6.append(row_data)
                    cellx = sheet1.cell(row=r,column=7)
                    cellx.value = int(row_data[0].strip())
                    print(str(cellx.value) + ' - ' + teamname)
                    r = r + 1
                    flag = 1
    if flag == 0:
        cellx = sheet1.cell(row=r,column=7)
        cellx.value = 0
        r = r + 1
    
stats_table7 = []
totalreboundurl = 'https://www.teamrankings.com/ncaa-basketball/stat/personal-fouls-per-game'
raw_html = urllib.request.urlopen(totalreboundurl)
html = bs(raw_html,'html.parser')
r = 2
for teamname in teamlist:

    
    Rankings = html.find_all('table')[0].find_all('tr')
    #print('Test' +teamname)         
    row_marker = 0
    row_data=[] 
    flag = 0
    for row in html.find_all('table')[0].find_all('tr'):
            column_marker = 0
            columns = row.find_all('td')
            row_data=[]
            for column in columns:
                if column_marker <= 1:
                    row_data.append(column.get_text())
                #print(row_data[column_marker])
                column_marker += 1                        
            for j in range(len(row_data)):
                #print('Test2'+teamname + ' '+ row_data[j] + ' '+ str(j) + str(j==1) + str(row_data[j].strip() == teamname.strip()))
                if j==1 and row_data[j].strip() == teamname.strip():
                    stats_table7.append(row_data)
                    cellx = sheet1.cell(row=r,column=8)
                    cellx.value = int(row_data[0].strip())
                    print(str(cellx.value) + ' - ' + teamname)
                    r = r + 1
                    flag = 1
    if flag == 0:
        cellx = sheet1.cell(row=r,column=8)
        cellx.value = 0
        r = r + 1
    
stats_table8 = []
totalreboundurl = 'https://www.teamrankings.com/ncaa-basketball/stat/turnovers-per-game'
raw_html = urllib.request.urlopen(totalreboundurl)
html = bs(raw_html,'html.parser')
r = 2
for teamname in teamlist:
    Rankings = html.find_all('table')[0].find_all('tr')
    #print('Test' +teamname)         
    row_marker = 0
    row_data=[]    
    flag = 0
    for row in html.find_all('table')[0].find_all('tr'):
            column_marker = 0
            columns = row.find_all('td')
            row_data=[]
            for column in columns:
                if column_marker <= 1:
                    row_data.append(column.get_text())
                #print(row_data[column_marker])
                column_marker += 1                        
            for j in range(len(row_data)):
                if j==1 and row_data[j].strip() == teamname.strip():
                    stats_table8.append(row_data)
                    cellx = sheet1.cell(row=r,column=9)
                    cellx.value = int(row_data[0].strip())
                    print(str(cellx.value) + ' - ' + teamname)
                    r = r + 1
                    flag = 1
    if flag == 0:
        cellx = sheet1.cell(row=r,column=9)
        cellx.value = 0
        r = r + 1
    
stats_table9 = []
totalreboundurl = 'https://www.teamrankings.com/ncaa-basketball/stat/opponent-turnovers-per-game'
raw_html = urllib.request.urlopen(totalreboundurl)
html = bs(raw_html,'html.parser')
r = 2
for teamname in teamlist:

    
    Rankings = html.find_all('table')[0].find_all('tr')
    #print('Test' +teamname)         
    row_marker = 0
    row_data=[]   
    flag = 0
    for row in html.find_all('table')[0].find_all('tr'):
            column_marker = 0
            columns = row.find_all('td')
            row_data=[]
            for column in columns:
                if column_marker <= 1:
                    row_data.append(column.get_text())
                #print(row_data[column_marker])
                column_marker += 1                        
            for j in range(len(row_data)):
                #print('Test2'+teamname + ' '+ row_data[j] + ' '+ str(j) + str(j==1) + str(row_data[j].strip() == teamname.strip()))
                if j==1 and row_data[j].strip() == teamname.strip():
                    stats_table9.append(row_data)
                    cellx = sheet1.cell(row=r,column=10)
                    cellx.value = int(row_data[0].strip())
                    print(str(cellx.value) + ' - ' + teamname)
                    r = r + 1
                    flag = 1
    if flag == 0:
        cellx = sheet1.cell(row=r,column=10)
        cellx.value = 0
        r = r + 1

stats_table10 = []
totalreboundurl = 'https://www.teamrankings.com/ncaa-basketball/stat/three-point-pct'
raw_html = urllib.request.urlopen(totalreboundurl)
html = bs(raw_html,'html.parser')
r = 2
for teamname in teamlist:

    Rankings = html.find_all('table')[0].find_all('tr')
    #print('Test' +teamname)         
    row_marker = 0
    row_data=[] 
    flag = 0
    for row in html.find_all('table')[0].find_all('tr'):
            column_marker = 0
            columns = row.find_all('td')
            row_data=[]
            for column in columns:
                if column_marker <= 1:
                    row_data.append(column.get_text())
                #print(row_data[column_marker])
                column_marker += 1                        
            for j in range(len(row_data)):
                #print('Test2'+teamname + ' '+ row_data[j] + ' '+ str(j) + str(j==1) + str(row_data[j].strip() == teamname.strip()))
                if j==1 and row_data[j].strip() == teamname.strip():
                    stats_table10.append(row_data)
                    cellx = sheet1.cell(row=r,column=11)
                    cellx.value = int(row_data[0].strip())
                    print(str(cellx.value) + ' - ' + teamname)
                    r = r + 1
                    flag = 1
    if flag == 0:
        cellx = sheet1.cell(row=r,column=11)
        cellx.value = 0
        r = r + 1
    
stats_table11 = []
totalreboundurl = 'https://www.teamrankings.com/ncaa-basketball/stat/opponent-three-point-pct'
raw_html = urllib.request.urlopen(totalreboundurl)
html = bs(raw_html,'html.parser')
r = 2
for teamname in teamlist:

    Rankings = html.find_all('table')[0].find_all('tr')
    #print('Test' +teamname)         
    row_marker = 0
    row_data=[] 
    flag = 0
    for row in html.find_all('table')[0].find_all('tr'):
            column_marker = 0
            columns = row.find_all('td')
            row_data=[]
            for column in columns:
                if column_marker <= 1:
                    row_data.append(column.get_text())
                #print(row_data[column_marker])
                column_marker += 1                        
            for j in range(len(row_data)):
                #print('Test2'+teamname + ' '+ row_data[j] + ' '+ str(j) + str(j==1) + str(row_data[j].strip() == teamname.strip()))
                if j==1 and row_data[j].strip() == teamname.strip():
                    stats_table11.append(row_data)
                    cellx = sheet1.cell(row=r,column=12)
                    cellx.value = int(row_data[0].strip())
                    print(str(cellx.value) + ' - ' + teamname)
                    r = r + 1
                    flag = 1
    if flag == 0:
        cellx = sheet1.cell(row=r,column=12)
        cellx.value = 0
        r = r + 1
    
    
Statswb.save("D:\\upwork\\NCAA_Stats.xlsx")
Statswb.close

wb_obj = openpyxl.load_workbook("D:\\upwork\\NCAA_Stats.xlsx")
sheet_obj = wb_obj.active 

rno = 2
for teamname in teamlist:
    for col in range(2,13):
        try:
            val = int(sheet1.cell(row=rno,column=col).value)
        except TypeError:
            val = 0
        if val > 0 and val <= 40:
            my_green = openpyxl.styles.colors.Color(rgb='00008800')
            my_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_green)
            sheet_obj.cell(row=rno,column=col).fill = my_fill
        elif val > 40 and val <= 80:
            my_lightgr = openpyxl.styles.colors.Color(rgb='0000FF00')
            my_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_lightgr)
            sheet_obj.cell(row=rno,column=col).fill = my_fill
        elif val > 80 and val <= 120:
            my_vlightgr = openpyxl.styles.colors.Color(rgb='00339966')
            my_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_vlightgr)
            sheet_obj.cell(row=rno,column=col).fill = my_fill
        elif val > 120 and val <= 160:
            my_lighty = openpyxl.styles.colors.Color(rgb='00FFFF99')
            my_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_lighty)
            sheet_obj.cell(row=rno,column=col).fill = my_fill
        elif val > 160 and val <= 200:
            my_vlighty = openpyxl.styles.colors.Color(rgb='00FFFF00')
            my_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_vlighty)
            sheet_obj.cell(row=rno,column=col).fill = my_fill
        elif val > 200 and val <= 240:
            my_vlighty = openpyxl.styles.colors.Color(rgb='00FF9900')
            my_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_vlighty)
            sheet_obj.cell(row=rno,column=col).fill = my_fill
        elif val > 240 and val <= 280:
            my_vlighty = openpyxl.styles.colors.Color(rgb='00FF8080')
            my_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_vlighty)
            sheet_obj.cell(row=rno,column=col).fill = my_fill
        elif val > 280 and val <= 320:
            my_vlighty = openpyxl.styles.colors.Color(rgb='00993300')
            my_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_vlighty)
            sheet_obj.cell(row=rno,column=col).fill = my_fill
        elif val > 320:
            my_vlighty = openpyxl.styles.colors.Color(rgb='00FF0000')
            my_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_vlighty)
            sheet_obj.cell(row=rno,column=col).fill = my_fill
        elif val == 0:
            my_vlighty = openpyxl.styles.colors.Color(rgb='00C0C0C0')
            my_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_vlighty)
            sheet_obj.cell(row=rno,column=col).fill = my_fill
    rno = rno + 1
            
        
  
wb_obj.save("D:\\upwork\\NCAA_Stats.xlsx")
wb_obj.close
